import openpyxl
import customtkinter as ctk
from customtkinter import filedialog
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
import math
import shutil
import xlwings
from pathlib import Path
import os

mainCable = None;
calculatedPoles = []
scriptPath = os.path.abspath(__file__)
scriptDirPath = os.path.dirname(scriptPath)
documentsDirPath = os.path.join(os.path.expanduser("~"), "Documents")
root = ctk.CTk()
root.withdraw()

def ReadExcelData(file_path):
    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path)
    
    # Get the active sheet
    sheet = workbook.active
    
    # Get the number of rows and columns

    rows = sheet.max_row
    cols = sheet.max_column
    
    # Initialize the data table
    data_table = []

    # Iterate through rows, starting from row 1 and incrementing by 4
    for row_start in range(1, rows + 1, 4):
        # Initialize an object for the current range
        data_object = []
        
        # Iterate through columns
        for col in range(1, cols + 1):
            # Get values from the current range
            cell_value = [sheet.cell(row=row, column=col).value for row in range(row_start, row_start + 4)]
            
            # Add values to the object
            data_object.append(cell_value)
        
        # Add the object to the data table
        data_table.append(data_object)

    return data_table
#table of tables -> one table one pole with all equipment

def countLength(cordsA, cordsB):
    x1, y1 = cordsA
    x2, y2 = cordsB

    length = round(math.sqrt((x2 - x1)**2 + (y2 - y1)**2))

    return length
def countDeq(coordsA):
    global mainCable
    if (mainCable is None):
        return 90

    coordsB = mainCable

    vector1 = [b - a for a, b in zip(coordsA[0], coordsB[0])]
    vector2 = [b - a for a, b in zip(coordsA[1], coordsB[1])]

    dotProduct = sum(a * b for a, b in zip(vector1, vector2))
    vectorLength1 = math.sqrt(sum(a**2 for a in vector1))
    vectorLength2 = math.sqrt(sum(b**2 for b in vector2))

    if vectorLength1 == 0 or vectorLength2 == 0:
        deg = 180
    else:
        degCos = dotProduct / (vectorLength1 * vectorLength2)
        deg = round(math.degrees(math.acos(min(1, max(-1, degCos)))))

    return deg + 90
def formatCablesString(inputString):
    #trim string
    trimmed_string = inputString.strip()

    #separate satring after +
    parts = trimmed_string.split('+')
    parts = trimmed_string.split('-')

    #format string
    for i in range(len(parts)):
        part = parts[i]

        if 'adss' in part:
            last_s_index = part.rfind('s')
            parts[i] = part[:last_s_index + 1] + ' ' + part[last_s_index + 1:]

        if 'al' in part:
            if 'l.' not in part:
                parts[i] = part.replace('l', 'l. ')
            else:
                parts[i] = part.replace('l.', 'l. ')

        if 'asxsn' in part:
            parts[i] = part.replace('n', 'n ')

    return parts
def formatCoordsString(coordsA):

    coordsA = coordsA.strip('()')
    coordsA = coordsA.split(' ')
    x1, y1 = map(float, coordsA[:2])
 
    return [x1, y1];
def getPoleFromData(data):
    formattedString = data.strip('()')
    separatedStrings = formattedString.split(' ')

    pole, number, function, station  = separatedStrings
    return {"pole":pole, "number": number, "function": function, "station" : station}
def handle_P(data, excel):
    indicator = data[0]
    #{"pole":pole, "function": function, "number": number}
    poleData = getPoleFromData(data[1])
    cords = data[3]
    
    print(poleData)
    excel['C78'] = poleData['number']; 
    excel['G78'] = poleData['pole'];
    excel['J78'] = poleData['function'].upper();
    excel['G88'] = 15; #mufa
    excel['L78'] = poleData['station'] 

    return poleData['station']
def handle_M(data, excel):
    global mainCable
    
    indicator = data[0].upper()
    cables = formatCablesString(data[1])
    coordsA = formatCoordsString(data[2])
    coordsB = formatCoordsString(data[3])

    rangeVectorA = [['E44', 'E45', 'E46', 'E47', 'E48', 'E49', 'E50', 'E51'],
                    ['H44', 'H45', 'H46', 'H47', 'H48', 'H49', 'H50', 'H51']]
    rangeVectorB = [['E52', 'E53', 'E54', 'E55', 'E56', 'E57', 'E58', 'E59'],
                    ['H52', 'H53', 'H54', 'H55', 'H56', 'H57', 'H58', 'H59']]
    
    rangeVectorSecondary = [['E65', 'E66', 'E67', 'E68', 'E69', 'E70', 'E71', 'E72', 'E73', 'E74', 'E75'],
                            ['H65', 'H66', 'H67', 'H68', 'H69', 'H70', 'H71', 'H72', 'H73', 'H74', 'H75']]

    def putValueToCell(cable, rangeVector):
        for i in range(len(rangeVector[0])):
            cellA = rangeVector[0][i]
            cellB = rangeVector[1][i]

            if (excel[cellA].value == None):
                cells = excel[cellA : cellB]                
                for cell in cells[0]:
                    cell.value = cable[0]
                    cable.pop(0)

                return

    for i in range(len(cables)):
        cableType = cables[i]
        cable = [cableType, None, countDeq([coordsA, coordsB]), countLength(coordsA, coordsB)]
        if (mainCable is None and indicator == "M"):
            putValueToCell(cable, rangeVectorA)
        else:
            if (indicator== "M"):
                putValueToCell(cable, rangeVectorB)
            else: 
                putValueToCell(cable, rangeVectorSecondary)

    #if this is first cable 
    if (mainCable is None and indicator == "M"):
        mainCable = [coordsA, coordsB]

def exportDataFromCalculatedExcel(excel):

    global calculatedPoles
    excel.active = excel["KALKULATOR"];
    sheet = excel.active;
    lp = len(calculatedPoles) + 1
    station = sheet['L78'].value.upper()
    number = sheet['C78'].value
    pole = sheet['G78'].value.upper()
    function = sheet['J78'].value.upper()
    
    sheetRange = [['C44', 'C45', 'C46', 'C47', 'C48', 'C49', 'C50', 'C51', 'C52', 'C53', 'C54', 'C55', 'C56', 'C57', 'C58', 'C59', 'C60', 'C61', 'C62', 'C63', 'C64', 'C65', 'C66', 'C67', 'C68', 'C69', 'C70', 'C71', 'C72', 'C73', 'C74', 'C75'],
                ['K44', 'K45', 'K46', 'K47', 'K48', 'K49', 'K50', 'K51', 'K52', 'K53', 'K54', 'K55', 'K56', 'K57', 'K58', 'K59', 'K60', 'K61', 'K62', 'K63', 'K64', 'K65', 'K66', 'K67', 'K68', 'K69', 'K70', 'K71', 'K72', 'K73', 'K74', 'K75']]
    
    dataFromSheetRange = []

    #for each row
    for i in range(len(sheetRange[0])):
        cellA = sheetRange[0][i]
        cellB = sheetRange[1][i]
        cells = sheet[cellA : cellB]                

        rowData = []
        for cell in cells[0]:
            rowData.append(cell.value)
        dataFromSheetRange.append(rowData)

    filterCell = 2 #3 column = cable type, if None == delete all array row
    filteredDataFromSheetRange = [row for row in dataFromSheetRange if row[filterCell] is not None]
    # filteredDataFromSheetRange = [[row for row in subRow if row is not None] for subRow in filteredDataFromSheetRange]
    filteredDataFromSheetRange = [[row for index, row in enumerate(eachRow) if index not in (1, 3)] for eachRow in filteredDataFromSheetRange]


    excel.active = excel[function]
    sheet = excel.active;
    
    maxX = round(float(sheet['D2'].value),2)
    maxY = round(float(sheet['D3'].value),2)
    realMaxX = round(maxX*0.1,2)
    realMaxY = round(maxY*0.1,2)
    calcX = round(float(sheet["B2"].value),2)
    calcY = round(float(sheet["B3"].value),2)
    addedX = round(float(sheet["G2"].value),2)
    addedY = round(float(sheet["G3"].value),2)

    pole = {
        "lp": lp,
        "station" : station,
        "number" : number,
        "pole" : pole,
        "function" : function, 
        "cables" : filteredDataFromSheetRange,
        "maxX" : maxX,
        "maxY" : maxY,
        "realMaxX": realMaxX,
        "realMaxY": realMaxY,
        "calcX" : calcX,
        "calcY" : calcY,
        "addedX" : addedX,
        "addedY" : addedY
    }

    calculatedPoles.append(pole)  
def cacheFormulasData(path, folderDir, number):

    # excel = win32.gencache.EnsureDispatch('Excel.Application')
    # excel.DisplayAlerts = False
    # workbook = excel.Workbooks.Open(rf'{path}')
    # # workbook.SaveAs(rf'{path}')
    # # workbook.Save(True, rf'{path}')
    # # workbook.Close(True, rf'{path}')
    # # workbook.Close()
    # excel.Quit()
    # del excel

    with xlwings.App(visible=False) as app:
        wb = xlwings.Book(path)
        wb.save(f"{folderDir}\\SŁUP_{number}.xlsx")
        wb.close()
    
    return f"{folderDir}\\SŁUP_{number}.xlsx"
def handleExcelFile(excel, path):
    # sourceExcel = "C:\\Users\\BFS\\Documents\\kalkulator.xlsx"
    global scriptDirPath
    sourceExcel = os.path.join(scriptDirPath, "kalkulator.xlsx")
    if (excel is None):
        shutil.copy(sourceExcel, path)
        excel = openpyxl.load_workbook(path)
        return excel
    else: 
        excel.save(path)
        excel.close()
def handleData(data, folderDir):

    for i in range(len(data)):
        global mainCable
        mainCable = None
        poleData = data[i]

        if (len(poleData) <=1): 
            print("error", poleData)
            return any
        
        
        newExcelPath = f'{folderDir}\\tempCalculatorFile.xlsx'
        excel = handleExcelFile(None, newExcelPath)
        excel.active = excel["KALKULATOR"];
        sheet = excel.active;
        poleStation = None
        for j in range(len(poleData)):
            table = poleData[j]
            if (table[0] is None):
                break
            indicator = table[0].upper()
            if (indicator == 'P'):
                poleStation = handle_P(table, sheet)
                next
            if (indicator == 'M' or indicator == "A"):
                handle_M(table, sheet)
                next
        
        handleExcelFile(excel, newExcelPath)

        def handleStationDir():
            if (poleStation is not None and os.path.exists(f'{folderDir}\\{poleStation}') is False):
                os.mkdir(f'{folderDir}\\{poleStation}')
            return f'{folderDir}\\{poleStation}'
        finalExcelPath = cacheFormulasData(newExcelPath, handleStationDir(), i)
        excel = openpyxl.load_workbook(finalExcelPath, data_only=True)

        exportDataFromCalculatedExcel(excel)
        handleExcelFile(excel, finalExcelPath)
    exportCalculatedData(folderDir)

    if (os.path.isfile(f'{folderDir}\\tempCalculatorFile.xlsx')):
        try:
            os.remove(f'{folderDir}\\tempCalculatorFile.xlsx')
        except:
            print("Wystąpił błąd podczas zamykania pliku tymczasowego excel, zgłoś problem developerowi")

def exportCalculatedData(folderDir):
    #check if file exist
    #open file
    global calculatedPoles
    global scriptDirPath
    # print(calculatedPoles[0]['cables'])
    # return
    path = f"{folderDir}\\Zestawienie obliczeń.xlsx"
    def handleExcelForExportedData():
        file = Path(path)
        if (file.exists()):
            return openpyxl.load_workbook(path)
        else: 
            # shutil.copy("C:\\Users\\BFS\\Documents\\Zestawienie obliczeń.xlsx", path)
            shutil.copy(os.path.join(scriptDirPath, "Zestawienie obliczeń.xlsx"), path)
            return openpyxl.load_workbook(path)
        
    excel = handleExcelForExportedData()
    sheet = excel.active;
    
    # excelRange = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W']
    excelCablesRange = ['F', 'G', 'G', 'H', 'I', 'J', 'K']
    excelRestCableDataRange = ['A','B','C','D','E', "L", 'M','N','O','P','Q','R','S','T',"U", "V"]
    
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    alignment = Alignment(horizontal='center', vertical='center')
    font = Font(bold=True)
    fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type = "solid")

    row = 3;
    for calcPole in calculatedPoles:
        lp = calcPole["lp"]
        station = calcPole["station"]
        number = calcPole["number"]
        pole = calcPole["pole"]
        function = calcPole["function"] 
        cables = calcPole["cables"]
        maxX = calcPole["maxX"]
        maxY = calcPole["maxY"]
        realMaxX = calcPole["realMaxX"]
        realMaxY = calcPole["realMaxY"]
        calcX = calcPole["calcX"]
        calcY = calcPole["calcY"]
        addedX = calcPole["addedX"]
        addedY = calcPole["addedY"]

        sheet[f'A{row}'].value = str(lp)
        sheet[f'B{row}'].value = station
        sheet[f'C{row}'].value = number
        sheet[f'D{row}'].value = pole
        sheet[f'E{row}'].value = function

        #catalog values
        sheet[f'L{row}'].value = str(maxX)
        sheet[f'M{row}'].value = str(maxY)

        #max * pole state
        sheet[f'N{row}'].value = str(realMaxX)
        sheet[f'O{row}'].value = str(realMaxY)

        #only electrical
        sheet[f'P{row}'].value = str(calcX)
        sheet[f'Q{row}'].value = str(calcY)

        sheet[f'R{row}'].value = function

        sheet[f'S{row}'].value = str(addedX)
        sheet[f'T{row}'].value = str(addedY)

        sheet[f'U{row}'].value = "Tak"
        sheet[f'V{row}'].value = "Dobry"

        cableRow = row

        
        for cable in cables:
            for i in range(len(cable)):

                
               
                
                cableData = cable[i]
                column = excelCablesRange[i]
                cell = sheet[f'{column}{cableRow}']

                if (i == 2): #degrees
                    continue

                if (i == 1): 
                    cell.fill = fill
                    cell.font = font     
                
                cell.value = cableData
                cell.alignment = alignment
                cell.border = border

             

            cableRow += 1

        cableRow -= 1
        for colRange in excelRestCableDataRange:
            cell = sheet[f'{colRange}{row}'];
            cell.border = border
            cell.alignment = alignment
            sheet.merge_cells(range_string = f'{colRange}{row}:{colRange}{cableRow}')
            # sheet.merge_cells(start_column = colRange ,start_row = row, end_column = colRange, end_row = cableRow)

        row = cableRow + 1

    handleExcelFile(excel, path)
# result_table = ReadExcelData("C:\\Users\\BFS\\Documents\\polesData.xlsx")\
tablePath = filedialog.askopenfilename(title="Wybierz plik Excel z zestawieniem słupów", filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))

if (bool (tablePath) is False):
    print("Nie wybrano ścieżki z zestawieniem słupów")
    print("Program zakończony niepowodzeniem")
    exit()

# result_table = ReadExcelData("C:\\Users\\BFS\\Documents\\polesData.xlsx")
result_table = ReadExcelData(tablePath)

def handleFolders():
    # resultsFolderPath = 'C:\\Users\\BFS\Documents\OBLICZENIA_WYTRZYMAŁOŚCI_SŁUPÓW'
    chosenResultsFolderPath = filedialog.askdirectory(title="Wybierz folder, w którym mają zostać stworzone obliczenia")
    if (bool(chosenResultsFolderPath is False)):
        print("Nie wybrano folderu do wstawienia obliczeń")
        return False
    resultsFolderPath = os.path.join(chosenResultsFolderPath, "OBLICZENIA_WYTRZYMAŁOŚCI_SŁUPÓW")

    if (os.path.exists(resultsFolderPath) is False):
        os.makedirs(resultsFolderPath)
    return resultsFolderPath


print("Program uruchomiony pomyślnie")    
folderDir = handleFolders()

if (folderDir is False):
    print("Program zakończony niepowodzeniem")
    exit()

handleData(result_table, folderDir)
print("Program zakończony powodzeniem")

